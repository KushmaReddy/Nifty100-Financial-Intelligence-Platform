import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment ,Font# -----------------------------
# Create Excel Workbook
# -----------------------------

workbook = Workbook()

# Remove default sheet
default_sheet = workbook.active
workbook.remove(default_sheet)

print("\nWorkbook created successfully!")
from pathlib import Path

# -----------------------------
# Paths
# -----------------------------

DB_PATH = Path("db/nifty100.db")
RAW_DATA_PATH = Path("data/raw")
OUTPUT_PATH = Path("outputs")

OUTPUT_PATH.mkdir(exist_ok=True)

# -----------------------------
# Connect Database
# -----------------------------

conn = sqlite3.connect(DB_PATH)

peer_percentiles = pd.read_sql(
    "SELECT * FROM peer_percentiles",
    conn
)

companies = pd.read_sql(
    "SELECT * FROM companies",
    conn
)

sectors = pd.read_sql(
    "SELECT * FROM sectors",
    conn
)

conn.close()

# -----------------------------
# Load Peer Groups
# -----------------------------

peer_groups = pd.read_excel(
    RAW_DATA_PATH / "peer_groups.xlsx"
)

print("Peer Percentiles :", peer_percentiles.shape)
print("Companies :", companies.shape)
print("Peer Groups :", peer_groups.shape)

print("\nPeer Groups Preview\n")
print(peer_groups.head())

# -----------------------------
# Merge Sector Information
# -----------------------------

peer_report = peer_percentiles.merge(
    sectors[
        [
            "company_id",
            "broad_sector",
            "sub_sector",
            "market_cap_category"
        ]
    ],
    on="company_id",
    how="left"
)

# -----------------------------
# Merge Benchmark Flag
# -----------------------------

peer_report = peer_report.merge(
    peer_groups[
        [
            "company_id",
            "is_benchmark"
        ]
    ],
    on="company_id",
    how="left"
)

print("\nPeer Report Shape :", peer_report.shape)

print("\nPeer Report Preview\n")

print(
    peer_report[
        [
            "company_id",
            "peer_group_name",
            "broad_sector",
            "sub_sector",
            "market_cap_category",
            "metric",
            "percentile_rank",
            "is_benchmark"
        ]
    ].head(15)
)
# -----------------------------
# Create One Sheet Per Peer Group
# -----------------------------

peer_groups_list = sorted(
    peer_report["peer_group_name"]
    .dropna()
    .unique()
)

print("\nPeer Groups Found:\n")

for group in peer_groups_list:

    workbook.create_sheet(
        title=group[:31]
    )

    print(group)

print(f"\nTotal Sheets Created : {len(workbook.sheetnames)}")

# -----------------------------
# Fill Each Sheet
# -----------------------------

for group in peer_groups_list:

    sheet = workbook[group[:31]]

    # Filter one peer group
    group_df = peer_report[
        peer_report["peer_group_name"] == group
    ].copy()

    # Convert long format to wide format
    report_df = (
        group_df.pivot_table(
            index="company_id",
            columns="metric",
            values="percentile_rank"
        )
        .reset_index()
    )

    # Add benchmark flag
    benchmark = (
        group_df[
            ["company_id", "is_benchmark"]
        ]
        .drop_duplicates()
    )

    report_df = report_df.merge(
        benchmark,
        on="company_id",
        how="left"
    )

    # Write header
    sheet.append(report_df.columns.tolist())

    # Write rows
    for row in report_df.itertuples(index=False):
        sheet.append(list(row))

print("\nAll sheets populated successfully!")

# -----------------------------
# Format Excel Sheets
# -----------------------------

header_fill = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
)

header_font = Font(
    color="FFFFFF",
    bold=True
)
green_fill = PatternFill(
    fill_type="solid",
    fgColor="92D050"
)

yellow_fill = PatternFill(
    fill_type="solid",
    fgColor="FFD966"
)

red_fill = PatternFill(
    fill_type="solid",
    fgColor="FF9999"
)

benchmark_fill = PatternFill(
    fill_type="solid",
    fgColor="FFC000"
)

for sheet in workbook.worksheets:

    # Header
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row in sheet.iter_rows(min_row=2):

        # Highlight benchmark row
        if row[-1].value == True:
            for cell in row:
                cell.fill = benchmark_fill

        # Colour percentile values
        for cell in row[1:-1]:

            if isinstance(cell.value, (int, float)):

                if cell.value >= 75:
                    cell.fill = green_fill

                elif cell.value >= 25:
                    cell.fill = yellow_fill

                else:
                    cell.fill = red_fill

    # Auto-fit columns
    for column_cells in sheet.columns:

        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )

        sheet.column_dimensions[
            column_cells[0].column_letter
        ].width = length + 3

print("\nExcel formatting applied successfully!")

# -----------------------------
# Save Workbook
# -----------------------------

output_file = OUTPUT_PATH / "peer_comparison.xlsx"

workbook.save(output_file)

print(f"\nWorkbook saved successfully!\n{output_file}")